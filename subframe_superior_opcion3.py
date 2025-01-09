import os  # Asegúrate de importar os
import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
from datetime import datetime
import json
import openpyxl
from unidecode import unidecode  # Para eliminar acentos

def create_subframe_superior(parent):
    """Crea el subframe superior con datos de entrada."""
    subframe = tk.Frame(parent, bg="white", relief="solid", bd=1)

    # Variables de estado
    datos_guardados = tk.BooleanVar(value=False)

    # Título
    tk.Label(subframe, text="Datos del servicio técnico", bg="lightblue", fg="black", 
             font=('Helvetica', 14, 'bold')).grid(row=0, column=0, columnspan=2, sticky="w", padx=10, pady=5)

    # Referencia del Servicio Técnico
    tk.Label(subframe, text="Referencia del Servicio Técnico:", bg='white', fg='black').grid(row=1, column=0, sticky="w", padx=10, pady=5)
    current_year = datetime.now().year
    default_ref_servicio = f"P{current_year}/"
    entry_ref_servicio = tk.Entry(subframe)
    entry_ref_servicio.insert(0, default_ref_servicio)
    entry_ref_servicio.grid(row=1, column=1, sticky="ew", padx=10, pady=5)

    # Fecha del Servicio
    tk.Label(subframe, text="Fecha del Servicio:", bg='white', fg='black').grid(row=2, column=0, sticky="w", padx=10, pady=5)
    date_entry = DateEntry(
        subframe, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy'
    )
    date_entry.delete(0, "end")
    date_entry.grid(row=2, column=1, sticky="ew", padx=10, pady=5)

    # Supervisor
    tk.Label(subframe, text="Supervisor/a:", bg='white', fg='black').grid(row=3, column=0, sticky="w", padx=10, pady=5)
    supervisors = ["Miguel Embid Segura"]
    combo_supervisor = ttk.Combobox(subframe, values=supervisors)
    combo_supervisor.grid(row=3, column=1, sticky="ew", padx=10, pady=5)

    # Cliente
    tk.Label(subframe, text="Cliente:", bg='white', fg='black').grid(row=5, column=0, sticky="w", padx=10, pady=5)
    entry_cliente = tk.Entry(subframe)
    entry_cliente.grid(row=5, column=1, sticky="ew", padx=10, pady=5)

    # Dirección del Cliente
    tk.Label(subframe, text="Dirección Cliente:", bg='white', fg='black').grid(row=6, column=0, sticky="w", padx=10, pady=5)
    dir_cliente = tk.Text(subframe, height=4, width=30)
    dir_cliente.grid(row=6, column=1, sticky="ew", padx=10, pady=5)

    # Marca
    tk.Label(subframe, text="Marca:", bg='white', fg='black').grid(row=10, column=0, sticky="w", padx=5, pady=5)
    entry_marca = tk.Entry(subframe)
    entry_marca.grid(row=10, column=1, sticky="ew", padx=5, pady=5)

    # Modelo
    tk.Label(subframe, text="Modelo:", bg='white', fg='black').grid(row=11, column=0, sticky="w", padx=5, pady=5)
    entry_modelo = tk.Entry(subframe)
    entry_modelo.grid(row=11, column=1, sticky="ew", padx=5, pady=5)

    # Número de Serie
    tk.Label(subframe, text="Número de serie:", bg='white', fg='black').grid(row=12, column=0, sticky="w", padx=5, pady=5)
    entry_numserie = tk.Entry(subframe)
    entry_numserie.grid(row=12, column=1, sticky="ew", padx=5, pady=5)

    # Botones
    def guardar_datos():
        referencia = entry_ref_servicio.get().strip().replace("/", "_")
        if not referencia:
            messagebox.showerror("Error", "La referencia del servicio no puede estar vacía.")
            return

        carpeta = os.path.join(os.getcwd(), referencia)
        if not os.path.exists(carpeta):
            os.makedirs(carpeta)

        # Datos a guardar
        datos = {
            "referencia_servicio": entry_ref_servicio.get().strip(),
            "fecha_servicio": date_entry.get(),
            "supervisor": combo_supervisor.get(),
            "cliente": entry_cliente.get(),
            "direccion_cliente": dir_cliente.get("1.0", tk.END).strip(),
            "marca": entry_marca.get(),
            "modelo": entry_modelo.get(),
            "numero_serie": entry_numserie.get(),
        }

        # Guardar en JSON y Excel (.xlsx)
        json_path = os.path.join(carpeta, "DS1.json")
        with open(json_path, "w") as json_file:
            json.dump(datos, json_file, indent=4)

        xlsx_path = os.path.join(carpeta, "DS1.xlsx")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Datos"
        for idx, (key, value) in enumerate(datos.items(), start=1):
            sheet.cell(row=1, column=idx, value=key)
            sheet.cell(row=2, column=idx, value=value)
        workbook.save(xlsx_path)

        datos_guardados.set(True)
        messagebox.showinfo("Guardado", f"Los datos se han guardado en:\n{carpeta}")

        # Mostrar subframe_inferior_opcion3 en la parte superior derecha (fila 0, columna 1)
        from subframe_inferior_opcion3 import create_subframe_inferior
        subframe_inferior = create_subframe_inferior(parent)
        subframe_inferior.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")


    def limpiar_campos():
        referencia = entry_ref_servicio.get().strip().replace("/", "_")
        carpeta = os.path.join(os.getcwd(), referencia)

        # Si los datos ya están guardados, elimina los archivos asociados
        if datos_guardados.get():
            json_path = os.path.join(carpeta, f"{referencia}.json")
            xlsx_path = os.path.join(carpeta, f"{referencia}.xlsx")
            if os.path.exists(json_path):
                os.remove(json_path)
            if os.path.exists(xlsx_path):
                os.remove(xlsx_path)
            messagebox.showinfo("Limpieza", "Archivos asociados eliminados.")

        # Limpia los campos de entrada
        for widget in subframe.winfo_children():
            if isinstance(widget, tk.Entry):
                widget.delete(0, "end")
            elif isinstance(widget, tk.Text):
                widget.delete("1.0", tk.END)
        date_entry.delete(0, "end")
        combo_supervisor.set("")
        datos_guardados.set(False)

    tk.Button(
        subframe, 
        text="Guardar", 
        command=guardar_datos, 
        bg="green", 
        fg="white"
    ).grid(row=99, column=0, sticky="e", padx=10, pady=10)

    tk.Button(
        subframe, 
        text="Limpiar", 
        command=limpiar_campos, 
        bg="red", 
        fg="white"
    ).grid(row=99, column=1, sticky="w", padx=10, pady=10)

    return subframe

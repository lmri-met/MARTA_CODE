import tkinter as tk
import os  
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
from datetime import datetime
import json
import openpyxl
from unidecode import unidecode  # Para eliminar acentos
import shutil
global calibration_data, coef_atenuacion, coef_conversion

def cargar_datos_iniciales(entry_ref_servicio, date_entry, combo_supervisor, entry_cliente,
                           dir_cliente, entry_marca, entry_modelo, entry_numserie, combo_proc,
                           combo_electrom_ref, combo_elecmonit_ref, combo_barom_ref,
                           combo_temp_ref, combo_crono_ref, combo_colim_ref,
                           combo_rangelec, combo_fcdsino,
                           selected_magnitude_var, label_rate_text_var, label_unit_text_var,
                           selected_quality_var, selected_chamber_var,
                           coef_conversion_var, calibration_coefficient_var,
                           air_attenuation_factor_var, calibration_factor_var, factor_correccion_var):
    """Carga datos desde datos_servicios.json si existe."""
    carpeta = "AAA_ST_REGISTROS"
    json_path = os.path.join(carpeta, "datos_servicios.json")

    # Verifica que el archivo existe antes de cargar datos
    if os.path.exists(json_path):
        with open(json_path, "r", encoding="utf-8") as json_file:
            datos = json.load(json_file)

        # Cargar datos solo si no son duplicados
        def actualizar_si_diferente(widget, nuevo_valor):
            if isinstance(widget, tk.Entry):
                actual = widget.get()
                if actual != nuevo_valor:
                    widget.delete(0, "end")
                    widget.insert(0, nuevo_valor)
            elif isinstance(widget, tk.Text):
                actual = widget.get("1.0", "end-1c").strip()
                if actual != nuevo_valor:
                    widget.delete("1.0", "end")
                    widget.insert("1.0", nuevo_valor)
            elif isinstance(widget, ttk.Combobox):
                actual = widget.get()
                if actual != nuevo_valor:
                    widget.set(nuevo_valor)

        # Aplicar lógica de actualización para cada campo
        actualizar_si_diferente(entry_ref_servicio, datos.get("ReferenciaServicio", ""))
        actualizar_si_diferente(entry_cliente, datos.get("Cliente", ""))
        actualizar_si_diferente(dir_cliente, datos.get("DireccionCliente", ""))
        actualizar_si_diferente(entry_marca, datos.get("Marca", ""))
        actualizar_si_diferente(entry_modelo, datos.get("Modelo", ""))
        actualizar_si_diferente(entry_numserie, datos.get("NumeroSerie", ""))
        actualizar_si_diferente(combo_supervisor, datos.get("Supervisor", ""))
        actualizar_si_diferente(combo_rangelec, datos.get("RangoElectrometro", ""))
        actualizar_si_diferente(combo_fcdsino, datos.get("FactorCorreccionDistancia", ""))
        actualizar_si_diferente(combo_proc, datos.get("Procedimiento", ""))
        actualizar_si_diferente(combo_electrom_ref, datos.get("ElectrometroPrincipal", ""))
        actualizar_si_diferente(combo_elecmonit_ref, datos.get("ElectrometroMonitor", ""))
        actualizar_si_diferente(combo_barom_ref, datos.get("Barometro", ""))
        actualizar_si_diferente(combo_temp_ref, datos.get("Termometro", ""))
        actualizar_si_diferente(combo_crono_ref, datos.get("Cronometro", ""))
        actualizar_si_diferente(combo_colim_ref, datos.get("Colimador", ""))

        # Actualizar variables
        selected_magnitude_var.set(datos.get("Magnitud", "Selecciona una magnitud"))
        label_rate_text_var.set(datos.get("UnidadesTasa", "..."))
        label_unit_text_var.set(datos.get("Unidades", "..."))
        selected_quality_var.set(datos.get("Calidad", "Selecciona una calidad"))
        selected_chamber_var.set(datos.get("CamaraPatron", "Selecciona una cámara"))
        coef_conversion_var.set(datos.get("CoeficienteConversion", "Coeficiente de Conversión: ..."))
        calibration_coefficient_var.set(datos.get("CoeficienteCalibracion", "Coef. Calibración: ..."))
        air_attenuation_factor_var.set(datos.get("CoeficienteAtenuacion", "Factor de Atenuación del Aire: ..."))
        calibration_factor_var.set(datos.get("FactorCalibracion", "Factor de Calibración: ..."))
        factor_correccion_var.set(datos.get("FactorCorreccion", "1.000"))

        # Fecha (DateEntry)
        fecha_servicio = datos.get("FechaServicio", datetime.now().strftime("%d/%m/%Y"))
        if date_entry.get_date() != datetime.strptime(fecha_servicio, "%d/%m/%Y"):
            date_entry.set_date(datetime.strptime(fecha_servicio, "%d/%m/%Y"))
    else:
        print(f"No se encontró el archivo {json_path}. Los campos estarán vacíos.")

def create_datos_servicio_frame(parent):
    subframe = tk.Frame(parent, bg="white", relief="solid", bd=1)

    # Cargar los datos JSON una sola vez
    try:
        with open("calibration_data.json", "r", encoding="utf-8") as f:
            calibration_data = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        calibration_data = {}
        print(f"[DEBUG] Error al cargar calibration_data.json: {e}")

    try:
        with open("CoefAtenAire.json", "r", encoding="utf-8") as f:
            coef_atenuacion = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        coef_atenuacion = []
        print(f"[DEBUG] Error al cargar CoefAtenAire.json: {e}")

    # Título
    titulo_frame = tk.Frame(subframe, bg="white")
    titulo_frame.pack(fill="x", pady=5, padx=10)
    tk.Label(titulo_frame, text="Datos del servicio técnico", bg="lightblue", fg="black",
             font=('Helvetica', 14, 'bold')).pack(anchor="w")

    # Configuración del ancho de las etiquetas
    label_width = 25  # Ancho uniforme para todas las etiquetas

    # Fila combinada: Referencia del Servicio Técnico y Fecha del Servicio
    fila1_2 = tk.Frame(subframe, bg="white")
    fila1_2.pack(fill="x", pady=5, padx=10)

    # Referencia del Servicio Técnico
    tk.Label(fila1_2, text="Referencia del Servicio Técnico:", bg="white", width=label_width, anchor="w").pack(side="left")
    entry_ref_servicio = tk.Entry(fila1_2, width=30)
    entry_ref_servicio.pack(side="left", padx=5)
    #entry_ref_servicio.insert(0, f"P{datetime.now().year}/")

    # Separación entre columnas
    tk.Label(fila1_2, text="", bg="white", width=5).pack(side="left")  # Espaciado opcional

    # Fecha del Servicio
    tk.Label(fila1_2, text="Fecha del Servicio:", bg="white", width=label_width, anchor="w").pack(side="left")
    date_entry = DateEntry(fila1_2, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
    date_entry.pack(side="left", padx=5)

    # Fila 3: Supervisor
    tk.Label(fila1_2, text="", bg="white", width=5).pack(side="left")  # Separador entre columnas
    tk.Label(fila1_2, text="Supervisor/a:", bg="white", width=label_width, anchor="w").pack(side="left")
    combo_supervisor = ttk.Combobox(fila1_2, values=["Miguel Embid Segura"], width=27)
    combo_supervisor.pack(side="left", padx=5)

    # Fila combinada: Cliente y Dirección del Cliente
    fila4_5 = tk.Frame(subframe, bg="white")
    fila4_5.pack(fill="x", pady=5, padx=10)

    # Cliente
    tk.Label(fila4_5, text="Cliente:", bg="white", width=label_width, anchor="w").pack(side="left")
    entry_cliente = tk.Entry(fila4_5, width=30)
    entry_cliente.pack(side="left", padx=5)

    # Separación entre columnas
    tk.Label(fila4_5, text="", bg="white", width=5).pack(side="left")  # Espaciado opcional

    # Dirección Cliente
    tk.Label(fila4_5, text="Dirección Cliente:", bg="white", width=label_width, anchor="w").pack(side="left")
    dir_cliente = tk.Text(fila4_5, height=4, width=32)
    dir_cliente.pack(side="left", padx=5)

    # Fila combinada: Marca, Modelo, y Número de Serie
    fila6_7_8 = tk.Frame(subframe, bg="white")
    fila6_7_8.pack(fill="x", pady=5, padx=10)

    # Marca
    tk.Label(fila6_7_8, text="Marca:", bg="white", width=label_width, anchor="w").pack(side="left")
    entry_marca = tk.Entry(fila6_7_8, width=30)
    entry_marca.pack(side="left", padx=5)

    # Separación entre columnas
    tk.Label(fila6_7_8, text="", bg="white", width=5).pack(side="left")  # Espaciado opcional

    # Modelo
    tk.Label(fila6_7_8, text="Modelo:", bg="white", width=label_width, anchor="w").pack(side="left")
    entry_modelo = tk.Entry(fila6_7_8, width=30)
    entry_modelo.pack(side="left", padx=5)

    # Separación entre columnas
    tk.Label(fila6_7_8, text="", bg="white", width=5).pack(side="left")  # Espaciado opcional

    # Número de Serie
    tk.Label(fila6_7_8, text="Número de serie:", bg="white", width=label_width, anchor="w").pack(side="left")
    entry_numserie = tk.Entry(fila6_7_8, width=30)
    entry_numserie.pack(side="left", padx=5)
    
    # Variables globales
    selected_quality_var = tk.StringVar(value="Selecciona una calidad")
    selected_chamber_var = tk.StringVar(value="Selecciona una cámara")
    calibration_factor_var = tk.StringVar(value="Factor de Calibración: ...")
    calibration_coefficient_var = tk.StringVar(value="Coef. Calibración: ...")
    air_attenuation_factor_var = tk.StringVar(value="Factor de Atenuación del Aire: ...")

    # Título
    fila10 = tk.Frame(subframe, bg="white")
    fila10.pack(fill="x", pady=10, padx=10)
    tk.Label(fila10, text="Datos del Tipo de Calibración", bg="lightblue", fg="black", 
             font=("Helvetica", 14, "bold")).pack(anchor="w")

    fila11_combined = tk.Frame(subframe, bg="white")
    fila11_combined.pack(fill="x", padx=10, pady=10)  # Ocupa todo el ancho con `fill="x"`

    # Variables y lógica de actualización
    magnitudes = [
        "Equivalente de dosis direccional H´(0,07)",
        "Equivalente de dosis direccional H´(3)",
        "Equivalente de dosis ambiental H*(10)",
        "Exposición",
        "Kerma en aire",
        "Dosis absorbida en aire"
    ]

    selected_magnitude_var = tk.StringVar(value="Selecciona una magnitud")
    label_rate_text_var = tk.StringVar(value="...")
    label_unit_text_var = tk.StringVar(value="...")

    def update_units(*args):
        selected_magnitude = selected_magnitude_var.get()
        if selected_magnitude in ["Equivalente de dosis direccional H´(0,07)",
                                  "Equivalente de dosis direccional H´(3)",
                                  "Equivalente de dosis ambiental H*(10)"]:
            label_rate_text_var.set("Sv/h")
            label_unit_text_var.set("Sv")
        elif selected_magnitude == "Exposición":
            label_rate_text_var.set("R/h")
            label_unit_text_var.set("R")
        elif selected_magnitude in ["Kerma en aire", "Dosis absorbida en aire"]:
            label_rate_text_var.set("Gy/h")
            label_unit_text_var.set("Gy")
        else:
            label_rate_text_var.set("N/A")
            label_unit_text_var.set("N/A")

    selected_magnitude_var.trace_add("write", update_units)

    # Widgets alineados a la izquierda
    tk.Label(fila11_combined, text="Magnitud de Medida:", bg="white", width=label_width, anchor="w").pack(side="left", padx=5)
    tk.OptionMenu(fila11_combined, selected_magnitude_var, *magnitudes).pack(side="left", padx=5)

    tk.Label(fila11_combined, text="Unidades de Tasa:", bg="white", width=label_width, anchor="w").pack(side="left", padx=5)
    tk.Label(fila11_combined, textvariable=label_rate_text_var, bg="white", anchor="w", width=20).pack(side="left", padx=5)

    tk.Label(fila11_combined, text="Unidades:", bg="white", width=label_width, anchor="w").pack(side="left", padx=5)
    tk.Label(fila11_combined, textvariable=label_unit_text_var, bg="white", anchor="w", width=20).pack(side="left", padx=5)

    # Fila combinada para selección de calidad y coeficiente de conversión
    fila13_combined = tk.Frame(subframe, bg="white")
    fila13_combined.pack(fill="x", padx=10, pady=5)

    # Cargar los datos del JSON
    with open("CoefConversion.json", "r", encoding="utf-8") as f:
        coef_data = json.load(f)

    # Listar las opciones de calidad
    qualities = [entry["Calidad"] for entry in coef_data]

    # Variable para calidad y coeficiente de conversión
    selected_quality_var = tk.StringVar(value="Selecciona una calidad")
    coef_conversion_var = tk.StringVar(value="Coeficiente de Conversión de Kerma en aire (Ka) a Magnitud de medida: ...")

    # Función para actualizar el coeficiente de conversión
    def update_coef(*args):
        selected_magnitude = selected_magnitude_var.get()
        selected_quality = selected_quality_var.get()

        # Buscar el coeficiente de conversión correspondiente
        coef = "N/A"
        for entry in coef_data:
            if entry["Calidad"] == selected_quality and selected_magnitude in entry:
                coef = entry[selected_magnitude]
                break

        coef_conversion_var.set(f"Coeficiente de Conversión de Kerma en aire (Ka) a Magnitud de medida: {coef}")

    # Actualización dinámica con trace
    selected_quality_var.trace_add("write", update_coef)

    # ComboBox para seleccionar la calidad de radiación
    tk.Label(fila13_combined, text="Calidad", bg="white", width=label_width, anchor="w").pack(side="left", padx=5)
    tk.OptionMenu(fila13_combined, selected_quality_var, *qualities).pack(side="left", padx=5)

    # Etiqueta para mostrar el coeficiente de conversión
    tk.Label(fila13_combined, textvariable=coef_conversion_var, bg="white", anchor="w").pack(side="left", padx=5)

    # Fila combinada para la selección de Cámara Patrón y Coeficiente de Calibración
    fila14_combined = tk.Frame(subframe, bg="white")
    fila14_combined.pack(fill="x", padx=10, pady=5)

    # Opciones de cámaras
    cameras = ["NE 2575-ns557-IR14D/014", "NE 2575-ns506-IR14D/006"]

    # Función para actualizar los campos
    def update_chamber_fields(*args):
        selected_quality = selected_quality_var.get()
        selected_chamber = selected_chamber_var.get()

        # Actualizar coeficiente de calibración
        if selected_quality.startswith("N") or selected_quality.startswith("L"):
            if selected_chamber == "NE 2575-ns557-IR14D/014":
                calibration_coefficient_var.set("43700")
            elif selected_chamber == "NE 2575-ns506-IR14D/006":
                calibration_coefficient_var.set("43660")
        elif selected_quality.startswith("W"):
            if selected_chamber == "NE 2575-ns557-IR14D/014":
                calibration_coefficient_var.set("43080")
            elif selected_chamber == "NE 2575-ns506-IR14D/006":
                calibration_coefficient_var.set("43150")
        else:
            calibration_coefficient_var.set("Calidad no soportada.")

        # Actualizar factor de calibración
        factor = "N/A"
        if selected_quality in calibration_data and selected_chamber in calibration_data[selected_quality]:
            factor = calibration_data[selected_quality][selected_chamber]
        calibration_factor_var.set(f"Factor de Calibración: {factor}")

        # Actualizar coeficiente de atenuación del aire
        air_factor = "N/A"
        for entry in coef_atenuacion:
            if entry["Calidad"] == selected_quality:
                air_factor = entry["Coef. Aten. aire"]
                break
    #    air_attenuation_factor_var.set(f"Factor de Atenuación del Aire: {air_factor}")
        air_attenuation_factor_var.set(f"{air_factor}")

    # Cargar los datos de los archivos JSON
    with open("CoefAtenAire.json", "r", encoding="utf-8") as f:
        coef_atenuacion = json.load(f)

    with open("Corrected_Factors.json", "r", encoding="utf-8") as f:
        corrected_factors = json.load(f)

    # Trace para actualizar los campos cuando cambia la selección
    selected_chamber_var.trace_add("write", update_chamber_fields)

    # Fila para la selección de Cámara Patrón
    fila14_combined = tk.Frame(subframe, bg="white")
    fila14_combined.pack(fill="x", padx=10, pady=5)

    tk.Label(fila14_combined, text="Cámara Patrón:", bg="white", width=label_width, anchor="w").pack(side="left", padx=5)
    combo_chamber = tk.OptionMenu(fila14_combined, selected_chamber_var, *cameras)
    combo_chamber.pack(side="left", padx=5)

    # Fila combinada para Coeficiente de Calibración y Coeficiente de Atenuación del Aire
    fila_calibracion_atenuacion = tk.Frame(subframe, bg="white")
    fila_calibracion_atenuacion.pack(fill="x", padx=10, pady=5)

    # Coeficiente de Calibración
    tk.Label(fila_calibracion_atenuacion, text="Coeficiente de Calibración:", bg="white", width=label_width, anchor="w").pack(side="left", padx=5)
    tk.Label(fila_calibracion_atenuacion, textvariable=calibration_coefficient_var, bg="white", anchor="w").pack(side="left", padx=5)

    # Separador entre columnas
    tk.Label(fila_calibracion_atenuacion, text="", bg="white", width=5).pack(side="left")

    # Coeficiente de Atenuación del Aire
    tk.Label(fila_calibracion_atenuacion, text="Coeficiente de Atenuación del Aire:", bg="white", width=label_width, anchor="w").pack(side="left", padx=5)
    tk.Label(fila_calibracion_atenuacion, textvariable=air_attenuation_factor_var, bg="white", anchor="w").pack(side="left", padx=5)

    # Fila para Factor de Calibración
    fila17_calibracion = tk.Frame(subframe, bg="white")
    fila17_calibracion.pack(fill="x", padx=10, pady=5)

    tk.Label(fila17_calibracion, text="Factor de Calibración:", bg="white", width=label_width, anchor="w").pack(side="left", padx=5)
    tk.Label(fila17_calibracion, textvariable=calibration_factor_var, bg="white", anchor="w").pack(side="left", padx=5)

    # Asignar los trazadores correctamente
    selected_chamber_var.trace_add("write", update_chamber_fields)
    selected_quality_var.trace_add("write", update_chamber_fields)

    # Título
    fila18 = tk.Frame(subframe, bg="white")
    fila18.pack(fill="x", pady=10, padx=10)
    tk.Label(fila18, text="Medidas para el cálculo del factor de distancia", bg="lightblue", fg="black", 
            font=("Helvetica", 14, "bold")).pack(anchor="w")

    # Campo: Rango del electrómetro
    fila_rango = tk.Frame(subframe, bg="white")
    fila_rango.pack(fill="x", pady=5, padx=10)
    tk.Label(fila_rango, text="Rango del electrómetro:", bg="white", width=35, anchor="w").pack(side="left")
    combo_rangelec = ttk.Combobox(fila_rango, values=["LOW", "HIGH"], width=35)
    combo_rangelec.pack(side="left", padx=10)

    # Campo: Factor de corrección de distancia
    fila_fcd = tk.Frame(subframe, bg="white")
    fila_fcd.pack(fill="x", pady=5, padx=10)
    tk.Label(fila_fcd, text="Factor de corrección de distancia:", bg="white", width=35, anchor="w").pack(side="left")
    combo_fcdsino = ttk.Combobox(fila_fcd, values=["Sí", "No"], width=35)
    combo_fcdsino.pack(side="left", padx=10)

    # Campo: Factor de corrección debido al rango del electrómetro
    fila_factor = tk.Frame(subframe, bg="white")
    fila_factor.pack(fill="x", pady=5, padx=10)
    tk.Label(fila_factor, text="Factor de corrección debido al rango del electrómetro:", bg="white", width=35, anchor="w").pack(side="left")
    factor_correccion_var = tk.StringVar(value="1.000")  # Valor inicial por defecto
    tk.Label(fila_factor, textvariable=factor_correccion_var, bg="white").pack(side="left", padx=10)

    # Función para actualizar el factor de corrección según el rango del electrómetro seleccionado
    def actualizar_factor(event):
        """Actualizar el valor del factor de corrección según el rango seleccionado."""
        rango = combo_rangelec.get()
        if rango == "LOW":
            factor_correccion_var.set("1.001")
        elif rango == "HIGH":
            factor_correccion_var.set("1.002")
        else:
            factor_correccion_var.set("1.000")

    # Vincular eventos para asegurar la actualización dinámica
    combo_rangelec.bind("<<ComboboxSelected>>", actualizar_factor)

    tk.Label(subframe, text="", bg="white").pack(fill="x", pady=5)

    # Título
    fila21 = tk.Frame(subframe, bg="white")
    fila21.pack(fill="x", pady=10, padx=10)
    tk.Label(fila21, text="Procedimientos y Equipamiento usados", bg="lightblue", fg="black", 
            font=("Helvetica", 14, "bold")).pack(anchor="w")

    # Fila combinada: Procedimiento, Electrómetro de medida y Electrómetro Monitor
    fila_proc_elecmon = tk.Frame(subframe, bg="white")
    fila_proc_elecmon.pack(fill="x", pady=5, padx=10)
    tk.Label(fila_proc_elecmon, text="Procedimiento:", bg="white", width=15, anchor="w").pack(side="left")
    combo_proc = ttk.Combobox(fila_proc_elecmon, values=["P-LMRI-C-12", "P-LMRI-C-13", "P-LMRI-C-26"], width=15)
    combo_proc.pack(side="left", padx=5)
    tk.Label(fila_proc_elecmon, text="Electr. Ppal:", bg="white", width=15, anchor="w").pack(side="left")
    combo_electrom_ref = ttk.Combobox(fila_proc_elecmon, values=["IR14D-18", "IR14D-28", "IR14D-41"], width=15)
    combo_electrom_ref.pack(side="left", padx=5)
    tk.Label(fila_proc_elecmon, text="Electr. Monitor:", bg="white", width=15, anchor="w").pack(side="left")
    combo_elecmonit_ref = ttk.Combobox(fila_proc_elecmon, values=["IR14D-18", "IR14D-28", "IR14D-41"], width=15)
    combo_elecmonit_ref.pack(side="left", padx=5)

    # Fila combinada: Barómetro, Termómetro, Cronómetro y Colimador
    fila_barom_temp_crono_colim = tk.Frame(subframe, bg="white")
    fila_barom_temp_crono_colim.pack(fill="x", pady=5, padx=10)
    tk.Label(fila_barom_temp_crono_colim, text="Barómetro:", bg="white", width=15, anchor="w").pack(side="left")
    combo_barom_ref = ttk.Combobox(fila_barom_temp_crono_colim, values=["IR14D-09", "IR14D-30", "SIMU"], width=15)
    combo_barom_ref.pack(side="left", padx=5)
    tk.Label(fila_barom_temp_crono_colim, text="Termómetro:", bg="white", width=15, anchor="w").pack(side="left")
    combo_temp_ref = ttk.Combobox(fila_barom_temp_crono_colim, values=["IR14D-10", "IR14D-32", "SIMU"], width=15)
    combo_temp_ref.pack(side="left", padx=5)
    tk.Label(fila_barom_temp_crono_colim, text="Cronómetro:", bg="white", width=15, anchor="w").pack(side="left")
    combo_crono_ref = ttk.Combobox(fila_barom_temp_crono_colim, values=["IR14D-19", "IR14D-18", "IR14D-28"], width=15)
    combo_crono_ref.pack(side="left", padx=5)
    tk.Label(fila_barom_temp_crono_colim, text="Colimador:", bg="white", width=15, anchor="w").pack(side="left")
    combo_colim_ref = ttk.Combobox(fila_barom_temp_crono_colim, values=["B20", "B40", "B60", "B80", "B100"], width=15)
    combo_colim_ref.pack(side="left", padx=5)

    # Función para mostrar botones al actualizar el campo "Colimador"
    def mostrar_botones(event):
        botones_frame.pack(fill="x", pady=10, padx=10)

    def guardar_datos():
        """Guarda los datos en archivos JSON y XLSX dentro de una carpeta específica."""
        carpeta = "AAA_ST_REGISTROS"

        # Crear carpeta si no existe
        if not os.path.exists(carpeta):
            os.makedirs(carpeta)

        # Definir los datos a guardar
        datos = {
            "ReferenciaServicio": entry_ref_servicio.get(),
            "FechaServicio": date_entry.get(),
            "Supervisor": combo_supervisor.get(),
            "Cliente": entry_cliente.get(),
            "DireccionCliente": dir_cliente.get("1.0", "end-1c"),
            "Marca": entry_marca.get(),
            "Modelo": entry_modelo.get(),
            "NumeroSerie": entry_numserie.get(),
            "Procedimiento": combo_proc.get(),
            "ElectrometroPrincipal": combo_electrom_ref.get(),
            "ElectrometroMonitor": combo_elecmonit_ref.get(),
            "Barometro": combo_barom_ref.get(),
            "Termometro": combo_temp_ref.get(),
            "Cronometro": combo_crono_ref.get(),
            "Colimador": combo_colim_ref.get(),
            "Magnitud": selected_magnitude_var.get(),
            "UnidadesTasa": label_rate_text_var.get(),
            "Unidades": label_unit_text_var.get(),
            "Calidad": selected_quality_var.get(),
            "CamaraPatron": selected_chamber_var.get(),
            "CoeficienteConversion": coef_conversion_var.get(),
            "CoeficienteCalibracion": calibration_coefficient_var.get(),
            "CoeficienteAtenuacion": air_attenuation_factor_var.get(),
            "FactorCalibracion": calibration_factor_var.get(),
            "FactorCorreccion": factor_correccion_var.get(),
            "RangoElectrometro": combo_rangelec.get(),
            "FactorCorreccionDistancia": combo_fcdsino.get()
        }

        # Guardar en archivo JSON
        json_path = os.path.join(carpeta, f"datos_servicios.json")
        with open(json_path, "w", encoding="utf-8") as json_file:
            json.dump(datos, json_file, ensure_ascii=False, indent=4)

        # Crear archivo Excel
        xlsx_path = os.path.join(carpeta, f"datos_servicios.xlsx")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "DatosServicio"

        # Escribir encabezados y datos
        for idx, (campo, valor) in enumerate(datos.items(), start=1):
            sheet.cell(row=1, column=idx, value=campo)  # Encabezado
            sheet.cell(row=2, column=idx, value=valor)  # Valor

        workbook.save(xlsx_path)
        messagebox.showinfo("Guardado", f"Datos guardados en la carpeta: {carpeta}")


    def resetear_datos():
        """Resetea los campos y elimina los archivos creados."""
        referencia = entry_ref_servicio.get().replace("/", "_")
        carpeta = referencia

        # Resetear campos
        entry_ref_servicio.delete(0, "end")
        date_entry.set_date(datetime.now())
        combo_supervisor.set("")
        entry_cliente.delete(0, "end")
        dir_cliente.delete("1.0", "end")
        entry_marca.delete(0, "end")
        entry_modelo.delete(0, "end")
        entry_numserie.delete(0, "end")
        combo_proc.set("")
        combo_electrom_ref.set("")
        combo_elecmonit_ref.set("")
        combo_barom_ref.set("")
        combo_temp_ref.set("")
        combo_crono_ref.set("")
        combo_colim_ref.set("")
        combo_rangelec.set(""), 
        combo_fcdsino.set(""),
        selected_magnitude_var.set("Selecciona una magnitud")
        label_rate_text_var.set("")
        label_unit_text_var.set("")
        selected_quality_var.set("Selecciona una calidad")
        selected_chamber_var.set("Selecciona una cámara")
        coef_conversion_var.set("Coeficiente de Conversión: ...")
        calibration_coefficient_var.set("Coef. Calibración: ...")
        air_attenuation_factor_var.set("Factor de Atenuación del Aire: ...")
        calibration_factor_var.set("Factor de Calibración: ...")
        factor_correccion_var.set("1.000")

        # Eliminar carpeta y archivos
        if os.path.exists(carpeta):
            shutil.rmtree(carpeta)
            messagebox.showinfo("Reset", f"Carpeta eliminada: {carpeta}")
        else:
            messagebox.showinfo("Reset", "No se encontró ninguna carpeta para eliminar.")

    # Crear marco para botones (visible desde el inicio)
    botones_frame = tk.Frame(subframe, bg="white")
    botones_frame.pack(fill="x", pady=10, padx=10)

    btn_guardar = tk.Button(botones_frame, text="Guardar", bg="green", fg="white", command=guardar_datos)
    btn_guardar.pack(side="left", padx=5)

    btn_reset = tk.Button(botones_frame, text="Resetear", bg="red", fg="white", command=resetear_datos)
    btn_reset.pack(side="left", padx=5)

    # Llamar a cargar_datos_iniciales una vez que todos los widgets han sido creados
    cargar_datos_iniciales(
        entry_ref_servicio, date_entry, combo_supervisor, entry_cliente,
        dir_cliente, entry_marca, entry_modelo, entry_numserie, combo_proc,
        combo_electrom_ref, combo_elecmonit_ref, combo_barom_ref,
        combo_temp_ref, combo_crono_ref, combo_colim_ref, combo_rangelec, combo_fcdsino,
        selected_magnitude_var, label_rate_text_var, label_unit_text_var,
        selected_quality_var, selected_chamber_var,
        coef_conversion_var, calibration_coefficient_var,
        air_attenuation_factor_var, calibration_factor_var, factor_correccion_var
    )

    return subframe

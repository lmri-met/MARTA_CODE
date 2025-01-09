import tkinter as tk
from tkinter import ttk, messagebox
import json
import os
import openpyxl

def create_subframe_inferior(parent):
    """Crear el subframe inferior con los datos solicitados para la Opción 3."""
    subframe = tk.Frame(parent, bg="white", relief="solid", bd=1)

    # Variable de estado para controlar si los datos están guardados
    datos_guardados = tk.BooleanVar(value=False)

    # Título
    tk.Label(subframe, text="Datos del Tipo de Calibración", bg="lightblue", fg="black", 
             font=("Helvetica", 14, "bold")).grid(row=0, column=0, columnspan=2, sticky="w", padx=10, pady=5)

    # Magnitud de Medida
    tk.Label(subframe, text="Magnitud de Medida:", bg="white", fg="black").grid(row=1, column=0, sticky="w", padx=10, pady=5)
    magnitudes = [
        "Equivalente de dosis direccional H´(0,07)",
        "Equivalente de dosis direccional H´(3)",
        "Equivalente de dosis ambiental H*(10)",
        "Exposición",
        "Kerma en aire",
        "Dosis absorbida en aire"
    ]
    combo_units = ttk.Combobox(subframe, values=magnitudes, width=35)
    combo_units.grid(row=1, column=1, sticky="ew", padx=10, pady=5)

    # Unidades de Tasa y Unidades
    label_rate = tk.Label(subframe, text="Unidades de Tasa: ........", bg="white", fg="black")
    label_rate.grid(row=2, column=0, sticky="w", padx=10, pady=5)
    label_unit = tk.Label(subframe, text="Unidades: ........", bg="white", fg="black")
    label_unit.grid(row=2, column=1, sticky="w", padx=10, pady=5)

    # Calidad de Radiación
    tk.Label(subframe, text="Calidad de Radiación:", bg="white", fg="black").grid(row=3, column=0, sticky="w", padx=10, pady=5)
    combo_quality = ttk.Combobox(subframe, width=35)
    combo_quality["values"] = [
        "L-10 1 m", "L-10 2,5 m", "L-20 1 m", "L-20 2,5 m",
        "L-30 1 m", "L-30 2,5 m", "L-35 1 m", "L-35 2,5 m",
        "L-55", "L-70", "L-100", "L-125", "L-170", "L-210", "L-240",
        "N-10 1 m", "N-10 2,5 m", "N-15 1 m", "N-15 2,5 m",
        "N-20 1 m", "N-20 2,5 m", "N-25 1 m", "N-25 2,5 m",
        "N-30 1 m", "N-30 2,5 m", "N-40 1 m", "N-40 2,5 m",
        "N-60", "N-80", "N-100", "N-120", "N-150", "N-200", "N-250", "N-300",
        "W-30 1 m", "W-30 2,5 m", "W-40 1 m", "W-40 2,5 m",
        "W-60", "W-80", "W-110", "W-150", "W-200", "W-250", "W-300"
    ]
    combo_quality.grid(row=3, column=1, sticky="ew", padx=10, pady=5)


    # Coef. Conversión de Ka a Magnitud Medida
    label_coef = tk.Label(subframe, text="Coef. Conversión: Seleccione opciones", bg="white", fg="black")
    label_coef.grid(row=4, column=0, columnspan=2, sticky="w", padx=10, pady=5)

    # Cámara Patrón
    tk.Label(subframe, text="Cámara Patrón:", bg="white", fg="black").grid(row=5, column=0, sticky="w", padx=10, pady=5)
    combo_chamber = ttk.Combobox(subframe, values=["NE 2575-ns557-IR14D/014", "NE 2575-ns506-IR14D/006"], width=35)
    combo_chamber.grid(row=5, column=1, sticky="ew", padx=10, pady=5)

    # Coeficiente de Calibración
    calibration_coefficient_var = tk.StringVar(value="Coef. calibración: ...")
    tk.Label(subframe, text="Coef. Calibración:", bg="white", fg="black").grid(row=6, column=0, sticky="w", padx=10, pady=5)
    tk.Label(subframe, textvariable=calibration_coefficient_var, bg="white", fg="black").grid(row=6, column=1, sticky="w", padx=10, pady=5)

    # Factor de Calibración
    calibration_factor_var = tk.StringVar(value="Factor de calibración: ...")
    tk.Label(subframe, text="Factor de Calibración:", bg="white", fg="black").grid(row=7, column=0, sticky="w", padx=10, pady=5)
    tk.Label(subframe, textvariable=calibration_factor_var, bg="white", fg="black").grid(row=7, column=1, sticky="w", padx=10, pady=5)

    # Factor de Atenuación del Aire
    fcaa_var = tk.StringVar(value="Factor de atenuación del aire: ...")
    tk.Label(subframe, text="Factor de Atenuación del Aire:", bg="white", fg="black").grid(row=8, column=0, sticky="w", padx=10, pady=5)
    tk.Label(subframe, textvariable=fcaa_var, bg="white", fg="black").grid(row=8, column=1, sticky="w", padx=10, pady=5)

    # Lógica para actualizar todos los valores
    def update_all_values(event=None):
        """Actualizar todos los valores dinámicamente."""
        selected_magnitude = combo_units.get()
        selected_quality = combo_quality.get()
        selected_chamber = combo_chamber.get()

        # Actualizar Unidades de Tasa y Unidades
        if selected_magnitude in ["Equivalente de dosis direccional H´(0,07)", 
                                  "Equivalente de dosis direccional H´(3)",
                                  "Equivalente de dosis ambiental H*(10)"]:
            label_rate.config(text="Unidades de Tasa: Sv/h")
            label_unit.config(text="Unidades: Sv")
        elif selected_magnitude == "Exposición":
            label_rate.config(text="Unidades de Tasa: R/h")
            label_unit.config(text="Unidades: R")
        elif selected_magnitude in ["Kerma en aire", "Dosis absorbida en aire"]:
            label_rate.config(text="Unidades de Tasa: Gy/h")
            label_unit.config(text="Unidades: Gy")
        else:
            label_rate.config(text="Unidades de Tasa: N/A")
            label_unit.config(text="Unidades: N/A")

        # Actualizar Coeficiente de Conversión
        entry = next(
            (item for item in coef_conversion if item.get("Calidad") == selected_quality), None
        )
        coef_value = entry.get(selected_magnitude, "Seleccione opciones") if entry else "Seleccione opciones"
        label_coef.config(text=f"Coef. Conversión: {coef_value}")

        # Actualizar Coeficiente de Calibración
        if selected_quality.startswith("N") or selected_quality.startswith("L"):
            if selected_chamber == "NE 2575-ns557-IR14D/014":
                calibration_coefficient_var.set("43700")
            elif selected_chamber == "NE 2575-ns506-IR14D/006":
                calibration_coefficient_var.set("43660")
        elif selected_quality.startswith("W"):
            if selected_chamber == "NE 2575-ns557-IR14D/014":
                calibration_coefficient_var.set("43080")
            elif selected_chamber == "NE 2575-ns506-IR14D/006":
                calibration_coefficient_var.set("43150")
        else:
            calibration_coefficient_var.set("Calidad no soportada.")

        # Actualizar Factor de Calibración
        calibration_entry = calibration_data.get(selected_quality, {}).get(selected_chamber, "N/A")
        calibration_factor_var.set(calibration_entry)

        # Actualizar Factor de Atenuación del Aire
        air_entry = next((item for item in coef_atenuacion if item.get("Calidad") == selected_quality), {})
        air_factor = air_entry.get("Coef. Aten. aire", "N/A")
        fcaa_var.set(air_factor)

    # Vincular eventos
    combo_units.bind("<<ComboboxSelected>>", update_all_values)
    combo_quality.bind("<<ComboboxSelected>>", update_all_values)
    combo_chamber.bind("<<ComboboxSelected>>", update_all_values)

    # Guardar datos
    def guardar_datos():
        referencia = entry_ref_servicio.get().strip().replace("/", "_")
        if not referencia:
            messagebox.showerror("Error", "La referencia del servicio no puede estar vacía.")
            return

        carpeta = os.path.join(os.getcwd(), referencia)
        if not os.path.exists(carpeta):
            os.makedirs(carpeta)

        # Datos con nombres de campo contraídos y sin acentos
        datos = {
            "referencia_servicio": entry_ref_servicio.get().strip(),
            "fecha_servicio": date_entry.get(),
            "supervisor": combo_supervisor.get(),
            "cliente": entry_cliente.get(),
            "direccion_cliente": dir_cliente.get("1.0", tk.END).strip(),
            "marca": entry_marca.get(),
            "modelo": entry_modelo.get(),
            "numero_serie": entry_numserie.get(),
        }

        # Eliminar acentos de los nombres de los campos
        datos = {unidecode(k): v for k, v in datos.items()}

        # Guardar en JSON
        json_path = os.path.join(carpeta, "DS1.json")
        with open(json_path, "w") as json_file:
            json.dump(datos, json_file, indent=4)

        # Guardar en Excel (.xlsx)
        xlsx_path = os.path.join(carpeta, "DS1.xlsx")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Datos"
        for idx, (key, value) in enumerate(datos.items(), start=1):
            sheet.cell(row=1, column=idx, value=key)
            sheet.cell(row=2, column=idx, value=value)
        workbook.save(xlsx_path)

        datos_guardados.set(True)
        messagebox.showinfo("Guardado", f"Los datos se han guardado en:\n{carpeta}")

        # Mostrar subframe_inferior_opcion3 en la parte superior derecha (fila 0, columna 1)
        from subframe_inferior_opcion3 import create_subframe_inferior
        subframe_inferior = create_subframe_inferior(parent)
        subframe_inferior.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")

    # Limpiar campos
    def limpiar_campos():
        if datos_guardados.get():
            referencia = f"calibracion_{combo_units.get().replace(' ', '_').lower()}_{combo_quality.get().replace(' ', '_').lower()}"
            referencia = referencia.replace("/", "_")
            carpeta = os.path.join(os.getcwd(), referencia)

            json_path = os.path.join(carpeta, f"{referencia}.json")
            xlsx_path = os.path.join(carpeta, f"{referencia}.xlsx")

            if os.path.exists(json_path):
                os.remove(json_path)
            if os.path.exists(xlsx_path):
                os.remove(xlsx_path)
            messagebox.showinfo("Limpieza", "Archivos eliminados.")

        combo_units.set("")
        combo_quality.set("")
        combo_chamber.set("")
        entry_atenuacion.delete(0, "end")
        entry_coef_calibracion.delete(0, "end")
        datos_guardados.set(False)

    # Botones
    tk.Button(
        subframe, 
        text="Guardar", 
        command=guardar_datos, 
        bg="green", 
        fg="white"
    ).grid(row=99, column=0, sticky="e", padx=10, pady=10)

    tk.Button(
        subframe, 
        text="Limpiar", 
        command=limpiar_campos, 
        bg="red", 
        fg="white"
    ).grid(row=99, column=1, sticky="w", padx=10, pady=10)

    return subframe